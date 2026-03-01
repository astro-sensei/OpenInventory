import json
import csv
import os
from typing import List
from models import Product, Sale, Supplier

DATA_DIR = "data"
PRODUCTS_FILE = os.path.join(DATA_DIR, "produits.json")
SALES_FILE = os.path.join(DATA_DIR, "ventes.json")
SUPPLIERS_FILE = os.path.join(DATA_DIR, "fournisseurs.json")
LOW_STOCK_THRESHOLD = 5


class Storage:
    def __init__(self):
        os.makedirs(DATA_DIR, exist_ok=True)
        self.products: List[Product] = []
        self.sales: List[Sale] = []
        self.suppliers: List[Supplier] = []
        self.load()

    # --- persistence JSON ---
    def load(self):
        self.products = self._load_json(PRODUCTS_FILE, Product)
        self.sales = self._load_json(SALES_FILE, Sale)
        self.suppliers = self._load_json(SUPPLIERS_FILE, Supplier)

    def save(self):
        self._save_json(PRODUCTS_FILE, self.products)
        self._save_json(SALES_FILE, self.sales)
        self._save_json(SUPPLIERS_FILE, self.suppliers)

    @staticmethod
    def _load_json(path, cls):
        if not os.path.exists(path):
            return []
        with open(path, "r", encoding="utf-8") as f:
            return [cls.from_dict(d) for d in json.load(f)]

    @staticmethod
    def _save_json(path, items):
        with open(path, "w", encoding="utf-8") as f:
            json.dump([i.to_dict() for i in items], f, ensure_ascii=False, indent=2)

    # --- CSV export/import ---
    def export_products_csv(self, path):
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["name", "reference", "quantity", "price", "category"])
            w.writeheader()
            for p in self.products:
                w.writerow(p.to_dict())

    def import_products_csv(self, path):
        with open(path, "r", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                p = Product.from_dict(row)
                existing = self.find_product(p.reference)
                if existing:
                    existing.name, existing.quantity, existing.price, existing.category = (
                        p.name, p.quantity, p.price, p.category,
                    )
                else:
                    self.products.append(p)
        self.save()

    def export_sales_csv(self, path):
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["reference", "product_name", "quantity", "unit_price", "date", "category"])
            w.writeheader()
            for s in self.sales:
                w.writerow(s.to_dict())

    # --- Excel export/import ---
    def export_products_excel(self, path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Produits"
        ws.append(["Nom", "Référence", "Quantité", "Prix", "Catégorie"])
        for p in self.products:
            ws.append([p.name, p.reference, p.quantity, p.price, p.category])
        ws2 = wb.create_sheet("Ventes")
        ws2.append(["Référence", "Produit", "Quantité", "Prix unitaire", "Date", "Catégorie"])
        for s in self.sales:
            ws2.append([s.reference, s.product_name, s.quantity, s.unit_price, s.date, s.category])
        ws3 = wb.create_sheet("Fournisseurs")
        ws3.append(["Nom", "Contact", "Produits fournis"])
        for sup in self.suppliers:
            ws3.append([sup.name, sup.contact, ", ".join(sup.products)])
        wb.save(path)

    def import_products_excel(self, path):
        from openpyxl import load_workbook
        wb = load_workbook(path)
        if "Produits" in wb.sheetnames:
            ws = wb["Produits"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                p = Product(name=str(row[0]), reference=str(row[1]),
                            quantity=int(row[2]), price=float(row[3]),
                            category=str(row[4]))
                existing = self.find_product(p.reference)
                if existing:
                    existing.name, existing.quantity, existing.price, existing.category = (
                        p.name, p.quantity, p.price, p.category,
                    )
                else:
                    self.products.append(p)
        self.save()

    # --- product operations ---
    def find_product(self, ref) -> Product | None:
        for p in self.products:
            if p.reference == ref:
                return p
        return None

    def search_products(self, query: str) -> List[Product]:
        q = query.lower()
        return [p for p in self.products if q in p.name.lower() or q in p.reference.lower()]

    def add_product(self, p: Product):
        if self.find_product(p.reference):
            raise ValueError(f"La référence '{p.reference}' existe déjà.")
        self.products.append(p)
        self.save()

    def update_product(self, ref, **kwargs):
        p = self.find_product(ref)
        if not p:
            raise KeyError("Produit non trouvé.")
        for k, v in kwargs.items():
            setattr(p, k, v)
        self.save()

    def delete_product(self, ref):
        p = self.find_product(ref)
        if not p:
            raise KeyError("Produit non trouvé.")
        self.products.remove(p)
        self.save()

    # --- sales ---
    def record_sale(self, ref: str, qty: int, date_str: str):
        p = self.find_product(ref)
        if not p:
            raise KeyError("Produit non trouvé.")
        if qty > p.quantity:
            raise ValueError(f"Stock insuffisant ({p.quantity} disponible(s)).")
        sale = Sale(
            reference=ref,
            product_name=p.name,
            quantity=qty,
            unit_price=p.price,
            date=date_str,
            category=p.category,
        )
        p.quantity -= qty
        self.sales.append(sale)
        self.save()
        return sale

    # --- reports ---
    def low_stock(self):
        return [p for p in self.products if p.quantity <= LOW_STOCK_THRESHOLD]

    def products_sorted(self, by="category"):
        key = (lambda p: p.category) if by == "category" else (lambda p: p.quantity)
        return sorted(self.products, key=key)

    def top_sold(self, start=None, end=None, limit=10):
        from collections import Counter
        c = Counter()
        for s in self.sales:
            if start and s.date < start:
                continue
            if end and s.date > end:
                continue
            c[s.product_name] += s.quantity
        return c.most_common(limit)

    def revenue_by_category(self, start=None, end=None):
        from collections import defaultdict
        rev = defaultdict(float)
        for s in self.sales:
            if start and s.date < start:
                continue
            if end and s.date > end:
                continue
            rev[s.category] += s.total
        return dict(rev)

    def total_revenue(self):
        return sum(s.total for s in self.sales)

    def stock_extremes(self):
        if not self.products:
            return None, None
        return min(self.products, key=lambda p: p.quantity), max(self.products, key=lambda p: p.quantity)

    # --- suppliers ---
    def add_supplier(self, s: Supplier):
        self.suppliers.append(s)
        self.save()

    def delete_supplier(self, name):
        self.suppliers = [s for s in self.suppliers if s.name != name]
        self.save()

    def update_supplier(self, old_name, **kwargs):
        for s in self.suppliers:
            if s.name == old_name:
                for k, v in kwargs.items():
                    setattr(s, k, v)
                self.save()
                return
        raise KeyError("Fournisseur non trouvé.")
